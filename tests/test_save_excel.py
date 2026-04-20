"""Unit tests for save_excel_streamlit — validate_data() and convert_df_to_excel()."""

from datetime import datetime
from io import BytesIO

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from sections.helpers.save_excel_streamlit import (
    ExcelExportError,
    convert_df_to_excel,
    validate_data,
)


class TestValidateData:
    def test_none_raises(self):
        with pytest.raises(ExcelExportError, match="None"):
            validate_data(None)

    def test_empty_dict_raises(self):
        with pytest.raises(ExcelExportError, match="empty"):
            validate_data({})

    def test_empty_dataframe_raises(self):
        with pytest.raises(ExcelExportError, match="empty"):
            validate_data(pd.DataFrame())

    def test_unsupported_type_raises(self):
        with pytest.raises(ExcelExportError):
            validate_data([1, 2, 3])

    def test_valid_dataframe_returned_unchanged_shape(self):
        df = pd.DataFrame({"a": [1, 2], "b": ["x", "y"]})
        result = validate_data(df)
        assert isinstance(result, pd.DataFrame)
        assert list(result.columns) == ["a", "b"]
        assert len(result) == 2

    def test_nan_replaced_with_empty_string(self):
        df = pd.DataFrame({"a": [1.0, np.nan], "b": ["ok", None]})
        result = validate_data(df)
        assert result["a"].iloc[1] == ""
        assert result["b"].iloc[1] == ""

    def test_dict_produces_key_value_dataframe(self):
        result = validate_data({"x": 1, "y": "hello"})
        assert list(result.columns) == ["Key", "Value"]
        assert set(result["Key"]) == {"x", "y"}

    def test_dict_none_value_becomes_empty_string(self):
        result = validate_data({"k": None})
        assert result["Value"].iloc[0] == ""

    def test_dict_datetime_value_formatted_as_string(self):
        dt = datetime(2024, 6, 15, 12, 30, 0)
        result = validate_data({"ts": dt})
        assert result["Value"].iloc[0] == "2024-06-15 12:30:00"

    def test_dict_nested_dict_converted_to_string(self):
        result = validate_data({"nested": {"a": 1}})
        val = result["Value"].iloc[0]
        assert isinstance(val, str)
        assert "a" in val

    def test_dict_list_value_converted_to_string(self):
        result = validate_data({"items": [1, 2, 3]})
        val = result["Value"].iloc[0]
        assert isinstance(val, str)

    def test_column_names_cast_to_string(self):
        df = pd.DataFrame({0: [1], 1: [2]})
        result = validate_data(df)
        assert all(isinstance(c, str) for c in result.columns)

    def test_numeric_key_in_dict_preserved_as_int(self):
        # int/float keys are kept as-is; only other types get str()-cast
        result = validate_data({42: "value"})
        assert result["Key"].iloc[0] == 42

    def test_original_dataframe_not_mutated(self):
        df = pd.DataFrame({"a": [np.nan]})
        original_val = df["a"].iloc[0]
        validate_data(df)
        assert pd.isna(df["a"].iloc[0]) == pd.isna(original_val)


class TestConvertDfToExcel:
    def test_returns_bytes(self):
        df = pd.DataFrame({"col1": [1, 2], "col2": ["a", "b"]})
        result = convert_df_to_excel(df)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_excel_is_valid_and_readable(self):
        df = pd.DataFrame({"x": [10, 20, 30]})
        result = convert_df_to_excel(df)
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        assert ws["A1"].value == "x"
        assert ws["A2"].value == 10

    def test_dict_input_produces_key_value_sheet(self):
        result = convert_df_to_excel({"key": "value", "num": 42})
        assert isinstance(result, bytes)
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        assert ws["A1"].value == "Key"

    def test_correct_row_count(self):
        df = pd.DataFrame({"a": range(5)})
        result = convert_df_to_excel(df)
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        assert ws.max_row == 6  # 1 header + 5 data rows

    def test_correct_column_count(self):
        df = pd.DataFrame({f"col_{i}": range(3) for i in range(10)})
        result = convert_df_to_excel(df)
        wb = load_workbook(BytesIO(result))
        assert wb.active.max_column == 10

    def test_none_input_raises(self):
        with pytest.raises(ExcelExportError):
            convert_df_to_excel(None)

    def test_sheet_named_sheet1(self):
        df = pd.DataFrame({"a": [1]})
        result = convert_df_to_excel(df)
        wb = load_workbook(BytesIO(result))
        assert "Sheet1" in wb.sheetnames

    def test_multirow_data_preserved(self):
        df = pd.DataFrame({"name": ["Alice", "Bob"], "score": [90, 85]})
        result = convert_df_to_excel(df)
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        names = [ws.cell(row=r, column=1).value for r in range(2, 4)]
        assert "Alice" in names
        assert "Bob" in names
